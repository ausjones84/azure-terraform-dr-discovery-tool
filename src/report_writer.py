"""
report_writer.py - Report Generation Engine
============================================
Generates Excel (.xlsx), Markdown (.md), and JSON (.json) reports
from discovery findings.

SAFETY: Read-only output generation. Never modifies Azure or Terraform.
"""

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from jinja2 import Environment, FileSystemLoader, select_autoescape

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from models import (
    DiscoveryReport, ComparisonStatus, RiskLevel,
    AzureResource, ComparisonResult
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants for Excel
# ---------------------------------------------------------------------------

HEADER_BG = "1F4E79"        # Dark blue
HEADER_FONT_COLOR = "FFFFFF"
RISK_COLORS = {
    RiskLevel.LOW.value: "C6EFCE",
    RiskLevel.MEDIUM.value: "FFEB9C",
    RiskLevel.HIGH.value: "FFC7CE",
    RiskLevel.CRITICAL.value: "FF0000",
}
STATUS_COLORS = {
    ComparisonStatus.TERRAFORM_MANAGED.value: "C6EFCE",
    ComparisonStatus.AZURE_ONLY.value: "FFC7CE",
    ComparisonStatus.TF_ONLY.value: "FFEB9C",
    ComparisonStatus.POSSIBLE_MATCH.value: "DDEEFF",
    ComparisonStatus.MODULE_AVAILABLE.value: "E2EFDA",
    ComparisonStatus.UNKNOWN.value: "F2F2F2",
}


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _header_style(ws, row: int, cols: List[str]):
    """Apply header row styling."""
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        cell.fill = PatternFill(
            start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 20


def _autofit_columns(ws, min_width=10, max_width=50):
    """Auto-size column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = adjusted


def _color_cell(cell, hex_color: str):
    """Apply background color to a cell."""
    cell.fill = PatternFill(
        start_color=hex_color, end_color=hex_color, fill_type="solid"
    )


# ---------------------------------------------------------------------------
# Excel report writer
# ---------------------------------------------------------------------------

def write_excel_report(report: DiscoveryReport, output_path: str) -> str:
    """
    Generate a multi-sheet Excel report.

    Sheets: Summary, Azure Resources, Private Endpoints, Networking,
    Terraform Matches, Drift Findings, Risks, Recommended Next Steps,
    Ticket Update Draft
    """
    if not EXCEL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel output. "
            "Install with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    ws_sum = wb.create_sheet("Summary")
    _write_summary_sheet(ws_sum, report)

    ws_az = wb.create_sheet("Azure Resources")
    _write_azure_resources_sheet(ws_az, report.azure_resources)

    ws_pe = wb.create_sheet("Private Endpoints")
    _write_private_endpoints_sheet(ws_pe, report.azure_resources)

    ws_net = wb.create_sheet("Networking")
    _write_networking_sheet(ws_net, report.azure_resources)

    ws_tf = wb.create_sheet("Terraform Matches")
    _write_terraform_matches_sheet(ws_tf, report.comparison_results)

    ws_drift = wb.create_sheet("Drift Findings")
    _write_drift_sheet(ws_drift, report)

    ws_risk = wb.create_sheet("Risks")
    _write_risks_sheet(ws_risk, report.comparison_results)

    ws_steps = wb.create_sheet("Recommended Next Steps")
    _write_next_steps_sheet(ws_steps, report)

    ws_ticket = wb.create_sheet("Ticket Update Draft")
    _write_ticket_sheet(ws_ticket, report)

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    logger.info("Excel report saved: %s", out_path)
    return str(out_path.resolve())


def _write_summary_sheet(ws, report: DiscoveryReport):
    data = [
        ("Tool", "Azure Terraform DR Discovery Tool"),
        ("Run Timestamp", report.run_timestamp),
        ("Service", report.service),
        ("Resource Name", report.resource_name),
        ("Subscription ID", report.subscription_id),
        ("Resource Group", report.resource_group),
        ("Repo Root", report.repo_root),
        ("Module Root", report.module_root),
        ("", ""),
        ("Resources Discovered", len(report.azure_resources)),
        ("Comparison Results", len(report.comparison_results)),
        ("Drift Entries Parsed", len(report.drift_entries)),
        ("", ""),
        ("Overall Status", report.summary_status.value if report.summary_status else ""),
        ("Overall Risk", report.overall_risk.value if report.overall_risk else ""),
        ("", ""),
        ("Key Findings", ""),
    ]
    for i, (label, value) in enumerate(data, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(value) if value is not None else "")
    row = len(data) + 1
    for finding in report.key_findings:
        ws.cell(row=row, column=2, value=finding)
        row += 1
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60


def _write_azure_resources_sheet(ws, resources: List[AzureResource]):
    headers = [
        "Name", "Resource Type", "Resource Group", "Subscription ID",
        "Location", "SKU", "Tags", "Identity Type",
        "Public Network Access", "Private Endpoints Count",
        "Diagnostic Settings Count", "Role Assignments Count",
    ]
    _header_style(ws, 1, headers)
    for row_idx, res in enumerate(resources, start=2):
        ws.cell(row=row_idx, column=1, value=res.name)
        ws.cell(row=row_idx, column=2, value=res.resource_type)
        ws.cell(row=row_idx, column=3, value=res.resource_group)
        ws.cell(row=row_idx, column=4, value=res.subscription_id)
        ws.cell(row=row_idx, column=5, value=res.location)
        ws.cell(row=row_idx, column=6, value=res.sku_name)
        tags_str = "; ".join(f"{k}={v}" for k, v in (res.tags or {}).items())
        ws.cell(row=row_idx, column=7, value=tags_str)
        identity_type = res.identity.type if res.identity else "None"
        ws.cell(row=row_idx, column=8, value=identity_type)
        pna = ws.cell(row=row_idx, column=9, value=res.public_network_access)
        if res.public_network_access and res.public_network_access.lower() == "enabled":
            _color_cell(pna, "FFC7CE")
        ws.cell(row=row_idx, column=10, value=len(res.private_endpoints))
        ws.cell(row=row_idx, column=11, value=len(res.diagnostic_settings))
        ws.cell(row=row_idx, column=12, value=len(res.role_assignments))
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _write_private_endpoints_sheet(ws, resources: List[AzureResource]):
    headers = [
        "Parent Resource", "PE Name", "Resource Group", "Location",
        "Connection State", "Private Link Service ID", "Group IDs",
        "VNet", "Subnet", "Private IP", "NIC Name", "DNS FQDNs",
    ]
    _header_style(ws, 1, headers)
    row_idx = 2
    for res in resources:
        for pe in res.private_endpoints:
            ws.cell(row=row_idx, column=1, value=res.name)
            ws.cell(row=row_idx, column=2, value=pe.name)
            ws.cell(row=row_idx, column=3, value=pe.resource_group)
            ws.cell(row=row_idx, column=4, value=pe.location)
            conn_cell = ws.cell(row=row_idx, column=5, value=pe.connection_state)
            if pe.connection_state and pe.connection_state.lower() != "approved":
                _color_cell(conn_cell, "FFC7CE")
            ws.cell(row=row_idx, column=6, value=pe.private_link_service_id)
            ws.cell(row=row_idx, column=7, value=", ".join(pe.group_ids))
            ws.cell(row=row_idx, column=8, value=pe.vnet_name)
            ws.cell(row=row_idx, column=9, value=pe.subnet_name)
            ip_cell = ws.cell(row=row_idx, column=10, value=pe.private_ip_address)
            if not pe.private_ip_address:
                _color_cell(ip_cell, "FFEB9C")
            ws.cell(row=row_idx, column=11, value=pe.nic_name)
            fqdns = "; ".join(d.fqdn for d in pe.dns_configs if d.fqdn)
            ws.cell(row=row_idx, column=12, value=fqdns)
            row_idx += 1
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _write_networking_sheet(ws, resources: List[AzureResource]):
    headers = [
        "Resource Name", "PE Name", "VNet", "Subnet", "Private IP",
        "NIC Name", "DNS FQDN", "DNS Private IP",
    ]
    _header_style(ws, 1, headers)
    row_idx = 2
    for res in resources:
        for pe in res.private_endpoints:
            dns_list = pe.dns_configs or [None]
            for dns in dns_list:
                ws.cell(row=row_idx, column=1, value=res.name)
                ws.cell(row=row_idx, column=2, value=pe.name)
                ws.cell(row=row_idx, column=3, value=pe.vnet_name)
                ws.cell(row=row_idx, column=4, value=pe.subnet_name)
                ws.cell(row=row_idx, column=5, value=pe.private_ip_address)
                ws.cell(row=row_idx, column=6, value=pe.nic_name)
                ws.cell(row=row_idx, column=7, value=dns.fqdn if dns else "")
                ws.cell(row=row_idx, column=8, value=dns.private_ip if dns else "")
                row_idx += 1
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _write_terraform_matches_sheet(ws, results: List[ComparisonResult]):
    headers = [
        "Resource Name", "File Path", "Line Number", "Match Type",
        "Confidence", "Confidence Level", "Matching Line",
    ]
    _header_style(ws, 1, headers)
    row_idx = 2
    for cr in results:
        if not cr.terraform_result:
            continue
        for match in cr.terraform_result.matches:
            ws.cell(row=row_idx, column=1, value=cr.azure_resource.name if cr.azure_resource else "")
            ws.cell(row=row_idx, column=2, value=match.file_path)
            ws.cell(row=row_idx, column=3, value=match.line_number)
            ws.cell(row=row_idx, column=4, value=match.match_type.value)
            ws.cell(row=row_idx, column=5, value=f"{match.confidence:.0%}")
            ws.cell(row=row_idx, column=6, value=match.confidence_label())
            ws.cell(row=row_idx, column=7, value=match.line_content[:200])
            row_idx += 1
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _write_drift_sheet(ws, report: DiscoveryReport):
    headers = [
        "Resource Name", "Resource Type", "Resource Group",
        "Subscription ID", "Status", "Notes",
    ]
    _header_style(ws, 1, headers)
    for row_idx, entry in enumerate(report.drift_entries, start=2):
        ws.cell(row=row_idx, column=1, value=entry.resource_name)
        ws.cell(row=row_idx, column=2, value=entry.resource_type)
        ws.cell(row=row_idx, column=3, value=entry.resource_group)
        ws.cell(row=row_idx, column=4, value=entry.subscription_id)
        ws.cell(row=row_idx, column=5, value=entry.status)
        ws.cell(row=row_idx, column=6, value=entry.notes)
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _write_risks_sheet(ws, results: List[ComparisonResult]):
    headers = [
        "Resource Name", "Status", "Risk Level", "Risk Notes",
        "Confidence", "Recommended Action",
    ]
    _header_style(ws, 1, headers)
    for row_idx, cr in enumerate(results, start=2):
        name = cr.azure_resource.name if cr.azure_resource else "N/A"
        ws.cell(row=row_idx, column=1, value=name)
        status_cell = ws.cell(row=row_idx, column=2, value=cr.status.value)
        _color_cell(status_cell, STATUS_COLORS.get(cr.status.value, "F2F2F2"))
        risk_cell = ws.cell(row=row_idx, column=3, value=cr.risk_level.value)
        _color_cell(risk_cell, RISK_COLORS.get(cr.risk_level.value, "F2F2F2"))
        ws.cell(row=row_idx, column=4, value="; ".join(cr.risk_notes))
        ws.cell(row=row_idx, column=5, value=f"{cr.confidence:.0%}")
        ws.cell(row=row_idx, column=6, value=cr.recommended_action)
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _write_next_steps_sheet(ws, report: DiscoveryReport):
    ws.cell(row=1, column=1, value="Recommended Next Steps").font = Font(bold=True, size=13)
    for row_idx, step in enumerate(report.next_steps, start=3):
        ws.cell(row=row_idx, column=1, value=f"{row_idx - 2}. {step}")
    ws.column_dimensions["A"].width = 100


def _write_ticket_sheet(ws, report: DiscoveryReport):
    ws.cell(row=1, column=1, value="Ticket Update Draft").font = Font(bold=True, size=13)
    ws.cell(row=3, column=1, value=report.ticket_update_text)
    ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[3].height = 200
    ws.cell(row=12, column=1, value="Teams Message Draft").font = Font(bold=True, size=13)
    ws.cell(row=14, column=1, value=report.teams_message_text)
    ws.cell(row=14, column=1).alignment = Alignment(wrap_text=True)


# ---------------------------------------------------------------------------
# Markdown report writer
# ---------------------------------------------------------------------------

def write_markdown_report(report: DiscoveryReport, output_path: str) -> str:
    """Generate a Markdown report from discovery findings."""
    lines = []
    ts = report.run_timestamp
    lines.append("# Azure DR Discovery Report")
    lines.append("")
    lines.append(f"**Generated:** {ts}  ")
    lines.append(f"**Service:** {report.service}  ")
    lines.append(f"**Resource:** {report.resource_name}  ")
    lines.append(f"**Subscription:** {report.subscription_id}  ")
    lines.append(f"**Overall Status:** `{report.summary_status.value}`  ")
    lines.append(f"**Overall Risk:** `{report.overall_risk.value}`  ")
    lines.append("")
    lines.append("## Key Findings")
    for finding in report.key_findings:
        lines.append(f"- {finding}")
    lines.append("")
    lines.append("## Azure Resources Discovered")
    for res in report.azure_resources:
        lines.append(f"### {res.name}")
        lines.append("| Field | Value |")
        lines.append("|-------|-------|")
        lines.append(f"| Resource Type | {res.resource_type} |")
        lines.append(f"| Resource Group | {res.resource_group} |")
        lines.append(f"| Location | {res.location} |")
        lines.append(f"| SKU | {res.sku_name} |")
        lines.append(f"| Identity | {res.identity.type if res.identity else 'None'} |")
        lines.append(f"| Public Network Access | {res.public_network_access} |")
        lines.append(f"| Private Endpoints | {len(res.private_endpoints)} |")
        lines.append("")
        if res.private_endpoints:
            lines.append("#### Private Endpoints")
            for pe in res.private_endpoints:
                lines.append(f"- **{pe.name}** | State: {pe.connection_state} | VNet: {pe.vnet_name} | Subnet: {pe.subnet_name} | IP: {pe.private_ip_address or '*Not set*'}")
            lines.append("")
    lines.append("## Terraform Search Results")
    for cr in report.comparison_results:
        name = cr.azure_resource.name if cr.azure_resource else "N/A"
        lines.append(f"### {name}")
        lines.append(f"- **Status:** `{cr.status.value}`")
        lines.append(f"- **Risk:** `{cr.risk_level.value}`")
        lines.append(f"- **Confidence:** {cr.confidence:.0%}")
        lines.append(f"- **Recommended Action:** {cr.recommended_action}")
        if cr.terraform_result and cr.terraform_result.matches:
            lines.append(f"- **Matches Found:** {len(cr.terraform_result.matches)}")
            lines.append("| File | Line | Match Type | Confidence |")
            lines.append("|------|------|------------|------------|")
            for m in cr.terraform_result.matches[:20]:
                lines.append(f"| {os.path.basename(m.file_path)} | {m.line_number} | {m.match_type.value} | {m.confidence:.0%} |")
        lines.append("")
    lines.append("## Recommended Next Steps")
    for i, step in enumerate(report.next_steps, start=1):
        lines.append(f"{i}. {step}")
    lines.append("")
    lines.append("## Ticket Update Draft")
    lines.append("")
    lines.append(report.ticket_update_text)
    lines.append("")
    lines.append("## Teams Message Draft")
    lines.append("")
    lines.append(report.teams_message_text)
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    logger.info("Markdown report saved: %s", out_path)
    return str(out_path.resolve())


# ---------------------------------------------------------------------------
# JSON report writer
# ---------------------------------------------------------------------------

def write_json_report(report: DiscoveryReport, output_path: str) -> str:
    """Generate a JSON report from discovery findings."""
    def _serialize(obj):
        if hasattr(obj, "__dataclass_fields__"):
            return {k: _serialize(v) for k, v in obj.__dict__.items()}
        if hasattr(obj, "value"):
            return obj.value
        if isinstance(obj, (list, tuple)):
            return [_serialize(i) for i in obj]
        if isinstance(obj, dict):
            return {k: _serialize(v) for k, v in obj.items()}
        return obj
    data = _serialize(report)
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    logger.info("JSON report saved: %s", out_path)
    return str(out_path.resolve())


# ---------------------------------------------------------------------------
# Template renderers
# ---------------------------------------------------------------------------

def render_ticket_update(report: DiscoveryReport, templates_dir: str) -> str:
    env = _get_jinja_env(templates_dir)
    try:
        tmpl = env.get_template("ticket_update.md.j2")
        return tmpl.render(report=report)
    except Exception as exc:
        logger.warning("Template render failed: %s. Using fallback.", exc)
        return _fallback_ticket_update(report)


def render_teams_message(report: DiscoveryReport, templates_dir: str) -> str:
    env = _get_jinja_env(templates_dir)
    try:
        tmpl = env.get_template("teams_message.md.j2")
        return tmpl.render(report=report)
    except Exception as exc:
        logger.warning("Template render failed: %s. Using fallback.", exc)
        return _fallback_teams_message(report)


def _get_jinja_env(templates_dir: str) -> Environment:
    path = Path(templates_dir)
    if not path.exists():
        path = Path(__file__).parent.parent / "templates"
    return Environment(
        loader=FileSystemLoader(str(path)),
        autoescape=select_autoescape([]),
        trim_blocks=True,
        lstrip_blocks=True,
    )


def _fallback_ticket_update(report: DiscoveryReport) -> str:
    resource_name = report.resource_name or (
        report.azure_resources[0].name if report.azure_resources else "unknown"
    )
    service = report.service.replace("_", " ").title()
    has_tf = any(
        cr.status.value == ComparisonStatus.TERRAFORM_MANAGED.value
        for cr in report.comparison_results
    )
    tf_status = (
        "Terraform deployment definition was located in the repository."
        if has_tf
        else f"Unable to confirm active Terraform deployment definition for "
             f"{resource_name} in the current repo search."
    )
    return (
        f"Completed initial review for {service}. "
        f"Validated Azure resource(s), private endpoint configuration, "
        f"network configuration, and Terraform module availability. "
        f"{tf_status} "
        f"No infrastructure changes were made. "
        f"Requesting confirmation on ownership/deployment path before proceeding."
    )


def _fallback_teams_message(report: DiscoveryReport) -> str:
    resource_name = report.resource_name or (
        report.azure_resources[0].name if report.azure_resources else "unknown"
    )
    service = report.service.replace("_", " ").title()
    return (
        f"Hi team, I completed the review for {service}. "
        f"I validated the existing Azure resource and reviewed the Terraform module/pattern. "
        f"I was not able to locate an active deployment definition in terraform-scripts "
        f"for {resource_name}. "
        f"Can you confirm whether this is managed through another repo, branch, or pipeline, "
        f"or if I should proceed with creating the Terraform deployment definition?"
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def write_all_reports(
    report: DiscoveryReport,
    output_dir: str,
    templates_dir: Optional[str] = None,
) -> dict:
    """Write all report formats (Excel, Markdown, JSON)."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = (report.resource_name or report.service or "discovery").replace(" ", "_")
    base_name = f"{safe_name}_{timestamp}"
    if templates_dir is None:
        templates_dir = str(Path(__file__).parent.parent / "templates")
    report.ticket_update_text = render_ticket_update(report, templates_dir)
    report.teams_message_text = render_teams_message(report, templates_dir)
    written = {}
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
    try:
        written["excel"] = write_excel_report(report, xlsx_path)
    except ImportError as exc:
        logger.warning("Excel output skipped: %s", exc)
    md_path = os.path.join(output_dir, f"{base_name}.md")
    written["markdown"] = write_markdown_report(report, md_path)
    json_path = os.path.join(output_dir, f"{base_name}.json")
    written["json"] = write_json_report(report, json_path)
    return written
